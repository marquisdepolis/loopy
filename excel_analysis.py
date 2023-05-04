# %%
import openpyxl
from openpyxl import load_workbook
from collections import defaultdict
import re
import os
import callgpt
import warnings
from tkinter import filedialog
from tkinter import *
import openai
warnings.filterwarnings("ignore")


def open_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as infile:
        return infile.read()


os.environ["OPENAI_API_KEY"] = open_file('Keys/openai_api_key.txt')
openai.api_key = open_file('Keys/openai_api_key.txt')
openai_api_key = openai.api_key
callgpt = callgpt.Ask()
MAX_ROW = 10


def load_workbook(file):
    if isinstance(file, str):
        return openpyxl.load_workbook(file, data_only=False)
    else:
        return file


def analyze_sheet_hierarchy(wb):
    sheet_hierarchy = "\nSheet hierarchy and structure:\n"
    # Implement any custom logic to identify sheet hierarchy or structure
    for sheet in wb:
        sheet_hierarchy += f" - {sheet.title}\n"
    return sheet_hierarchy


def analyze_named_ranges(wb):
    named_ranges = "\nNamed ranges:\n"
    for name, range_obj in wb.defined_names.items():
        if name is not None:
            named_ranges += f" - {name}: {range_obj}\n"
    return named_ranges


def analyze_sheet_contents(wb):
    sheet_contents = "\nSheet contents:\n"
    for sheet in wb:
        sheet_contents += f"\nSheet: {sheet.title}\n"
        sheet_contents += "Column names:\n"
        for cell in sheet[1]:
            column_letter = cell.coordinate.split('!')[-1][:-1]
            if cell.value is not None:
                sheet_contents += f" - {column_letter}: {cell.value}\n"
    return sheet_contents
    # Add logic to identify key cells or areas and input, calculation, and output areas


def analyze_formulae(wb):
    formulae = "\nFormulae:\n"
    for sheet in wb:
        formulae += f"\nSheet: {sheet.title}\n"
        formulae += "Formulas:\n"
        for row in sheet.iter_rows(min_row=2, max_row=MAX_ROW):
            for cell in row:
                if cell.data_type == "f":
                    if cell.value is not None:
                        formulae += f" - Cell {cell.coordinate}: {cell.value}\n"
    return formulae


def trace_dependencies(wb, sheet_name, cell, dependencies):
    formula = cell.value
    cell_references = re.findall(
        r"((?:(?:\')?[A-Za-z0-9\s]+(?:\')?\!)?\$?[A-Z]+\$?\d+)", formula)

    for ref in cell_references:
        if '!' in ref:
            ref_sheet_name, ref_cell_coord = ref.split('!')
            ref_sheet_name = ref_sheet_name.strip("'")
        else:
            ref_sheet_name = sheet_name
            ref_cell_coord = ref

        ref_cell = wb[ref_sheet_name][ref_cell_coord]

        if ref_cell.data_type == "f":
            full_ref = f"{ref_sheet_name}!{ref_cell_coord}"
            if full_ref not in dependencies[f"{sheet_name}!{cell.coordinate}"]:
                dependencies[f"{sheet_name}!{cell.coordinate}"].add(full_ref)
                trace_dependencies(wb, ref_sheet_name, ref_cell, dependencies)


def get_alpha(cell_reference):
    return re.sub(r'\d', '', cell_reference)


def get_numeric(cell_reference):
    return re.sub(r'[A-Za-z\$]', '', cell_reference)


def update_row_number(formula, offset):
    cell_references = re.findall(r"(\$?[A-Z]+\$?\d+)", formula)
    for ref in cell_references:
        alpha, numeric = get_alpha(ref), get_numeric(ref)
        new_numeric = int(numeric) + offset
        formula = formula.replace(ref, f"{alpha}{new_numeric}")
    return formula


def is_same_formula_row(row1, row2, match_percentage=0.5):
    matching_cells = 0
    total_cells = len(row1)

    for c1, c2 in zip(row1, row2):
        if c1.data_type == c2.data_type == "f":
            offset = c2.row - c1.row
            updated_formula = update_row_number(c1.value, offset)
            if updated_formula == c2.value:
                matching_cells += 1
        elif c1.value == c2.value:
            matching_cells += 1

    return matching_cells / total_cells >= match_percentage


def analyze_data_flows(wb):
    data_flows = "\nData flows and dependencies:\n"
    for sheet in wb:
        data_flows += f"\nSheet: {sheet.title}\n"
        dependencies = defaultdict(set)
        prev_row = None

        for row in sheet.iter_rows(min_row=1, max_row=MAX_ROW):
            if prev_row is not None and is_same_formula_row(prev_row, row):
                break

            for cell in row:
                if cell.data_type == "f":
                    trace_dependencies(wb, sheet.title, cell, dependencies)

            prev_row = row

        for cell_coord, dependent_cells in dependencies.items():
            data_flows += f" - {cell_coord}: {', '.join(dependent_cells)}\n"
    return data_flows


def analyze_external_links(wb):
    external_links = "\nExternal links:\n"
    if hasattr(wb, 'external_links') and wb.external_links:
        for link in wb.external_links:
            external_links += f" - {link.target}\n"
    return external_links


def analyze_key_outputs(wb):
    key_outputs = "\nKey outputs and their dependencies:\n"
    # Implement custom logic to identify key outputs and trace their dependencies
    return key_outputs


def main(file_path):
    wb = load_workbook(file_path)
    results = ""
    results += analyze_sheet_hierarchy(wb)
    results += analyze_named_ranges(wb)
    results += analyze_sheet_contents(wb)
    results += analyze_formulae(wb)
    results += analyze_data_flows(wb)
    results += analyze_external_links(wb)
    results += analyze_key_outputs(wb)
    print(results)
    return results
